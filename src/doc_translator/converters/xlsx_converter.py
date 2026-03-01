from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from typing import Any
import logging

logger = logging.getLogger(__name__)


def translate_xlsx(input_path: str, output_path: str, translator: Any, target_lang: str):
    """Translate an XLSX file using batch translation while preserving
    cell styles, formulas, merged cells, and sheet structure."""
    wb = load_workbook(filename=input_path)

    # --- collect every string cell that is writable ---
    cell_refs: list[Any] = []   # openpyxl Cell objects
    texts: list[str] = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if isinstance(val, str) and val.strip() and not val.startswith("="):
                    cell_refs.append(cell)
                    texts.append(val)

    logger.info("XLSX '%s': collected %d translatable string cells", input_path, len(texts))

    # --- batch-translate ---
    if texts:
        try:
            translated = translator.translate_batch(texts, target_lang)
        except Exception:
            logger.exception("Batch translation failed for XLSX; falling back to per-item")
            translated = []
            for t in texts:
                try:
                    r = translator.translate_text(t, target_lang)
                    translated.append(r if r is not None else t)
                except Exception:
                    logger.exception("Per-item fallback also failed")
                    translated.append(t)

        # --- write results back ---
        errors = 0
        for cell, orig, tr in zip(cell_refs, texts, translated):
            if tr is None:
                tr = orig
                errors += 1
            cell.value = tr
    else:
        logger.warning("No translatable text found in XLSX file '%s'", input_path)
        errors = 0

    wb.save(output_path)
    if errors:
        raise RuntimeError(f"Translation completed with {errors} failed cells")
