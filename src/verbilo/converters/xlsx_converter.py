from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, TextBlock
from typing import Any, Callable
import logging
import threading
import unicodedata
import re
from ..utils import CancelledError

logger = logging.getLogger(__name__)

# Control-character pattern: matches C0/C1 control chars except tab, newline, carriage return,
# plus invisible Unicode format/zero-width characters that can silently corrupt translation.
_CONTROL_CHAR_RE = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f'
    r'\u00ad'          # soft hyphen
    r'\u200b-\u200d'   # zero-width space / non-joiner / joiner
    r'\u2060'          # word joiner
    r'\ufeff]'         # BOM / zero-width no-break space
)

# Separator used to group multiple cells into a single translation unit.
# Chosen to be extremely unlikely in real spreadsheet data.
_CELL_SEP_TOKEN = "\u27EASEP\u27EB"
_CELL_SEP = f"\n{_CELL_SEP_TOKEN}\n"
_CELL_SEP_SPLIT_RE = re.compile(rf"\s*{re.escape(_CELL_SEP_TOKEN)}\s*")

# Maximum characters per grouped row before falling back to per-cell.
_GROUP_MAX_CHARS = 4000


def _sanitize_text(text: str) -> str:
    # Normalize Unicode and strip problematic control characters
    text = unicodedata.normalize("NFC", text)
    text = _CONTROL_CHAR_RE.sub("", text)
    return text


def _split_grouped_row_translation(text: str) -> list[str]:
    if _CELL_SEP in text:
        return text.split(_CELL_SEP)
    if _CELL_SEP_TOKEN not in text:
        return [text]
    return _CELL_SEP_SPLIT_RE.split(text)


def translate_xlsx(input_path: str, output_path: str, translator: Any, target_lang: str, *, cancel_event: threading.Event | None = None, source_lang: str = "auto", progress_callback: 'Callable[[int, int], None] | None' = None):
    # batch-translate XLSX with row-level contextual grouping.
    # When source_lang=="auto" row grouping is skipped so each cell is its own translation
    # unit, letting the API auto-detect the language per cell.
    wb = load_workbook(filename=input_path, rich_text=True)

    # --- collect every string cell that is writable, grouped by row ---
    # Each row group: list of (cell, sanitized_text) pairs
    # rich_segs: list of (cell, original_CellRichText, segment_index, sanitized_text)
    RowGroup = list[tuple[Any, str]]
    row_groups: list[RowGroup] = []
    current_row: RowGroup = []
    current_row_key: tuple | None = None  # (sheet_title, row_number)
    rich_segs: list[tuple[Any, CellRichText, int, str]] = []

    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=False):
            row_cells: RowGroup = []
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if isinstance(val, str) and not val.startswith("="):
                    sanitized = _sanitize_text(val)
                    if sanitized.strip():
                        row_cells.append((cell, sanitized))
                elif isinstance(val, CellRichText):
                    for seg_idx, seg in enumerate(val):
                        if isinstance(seg, TextBlock):
                            text = seg.text
                        elif isinstance(seg, str):
                            text = seg
                        else:
                            continue
                        if text and text.strip():
                            rich_segs.append((cell, val, seg_idx, _sanitize_text(text)))
            if row_cells:
                row_groups.append(row_cells)

    total_cells = sum(len(rg) for rg in row_groups)
    logger.info(
        "XLSX '%s': collected %d plain-text cells in %d rows, %d rich-text segments",
        input_path, total_cells, len(row_groups), len(rich_segs),
    )

    if not row_groups and not rich_segs:
        logger.warning("No translatable text found in XLSX file '%s'", input_path)
        wb.save(output_path)
        return

    # --- build translation units: group rows or send individually ---
    # A "unit" is either a joined row (multiple cells with separator) or a single cell.
    # After translation we split on separator to recover per-cell results.
    units: list[str] = []
    # Map: unit_index -> list of (cell, original_text) to write back
    unit_cells: list[list[tuple[Any, str]]] = []

    # In auto-detect mode do NOT group cells — every cell is its own unit so the
    # translation API sees a single-language segment and can auto-detect correctly.
    group_rows = source_lang != "auto"

    for rg in row_groups:
        row_texts = [t for _, t in rg]
        total_chars = sum(len(t) for t in row_texts)

        if group_rows and len(rg) > 1 and total_chars <= _GROUP_MAX_CHARS:
            # Group the row into a single unit with separators
            units.append(_CELL_SEP.join(row_texts))
            unit_cells.append(rg)
        else:
            # Send each cell individually (row too large, single cell, or auto mode)
            for cell, text in rg:
                units.append(text)
                unit_cells.append([(cell, text)])

    # --- batch-translate (plain cells + rich-text segments in one call) ---
    plain_count = len(units)
    all_units = units + [t for _, _, _, t in rich_segs]
    total_units = len(all_units)
    try:
        translated_all = translator.translate_batch(all_units, target_lang, cancel_event=cancel_event)
    except CancelledError:
        raise
    except Exception:
        logger.exception("Batch translation failed for XLSX; falling back to per-item")
        translated_all = []
        for t in all_units:
            try:
                r = translator.translate_text(t, target_lang)
                translated_all.append(r if r is not None else t)
            except Exception:
                logger.exception("Per-item fallback also failed")
                translated_all.append(t)

    if progress_callback is not None:
        progress_callback(total_units, total_units)

    plain_translated = translated_all[:plain_count]
    rich_translated = translated_all[plain_count:]

    # --- write results back (plain cells) ---
    errors = 0
    for cells_in_unit, tr_text in zip(unit_cells, plain_translated):
        if tr_text is None:
            # Translation returned None — keep originals
            errors += 1
            continue

        if len(cells_in_unit) == 1:
            # Single cell unit — direct assignment
            cell, orig = cells_in_unit[0]
            cell.value = tr_text
        else:
            # Grouped row — split on separator
            parts = _split_grouped_row_translation(tr_text)
            if len(parts) == len(cells_in_unit):
                for (cell, orig), part in zip(cells_in_unit, parts):
                    translated_part = part.strip() if part else orig
                    # If the part came back identical to the original the API likely
                    # skipped it (e.g. mixed-language row — Russian + Chinese grouped
                    # together; the API translated the dominant language and left the
                    # minority language segment untouched).  Retry as a standalone call
                    # so the API sees a clean single-language segment.
                    if translated_part == orig:
                        try:
                            r = translator.translate_text(orig, target_lang)
                            translated_part = r if r is not None else orig
                        except Exception:
                            logger.exception("Per-cell retry failed for unchanged segment")
                    cell.value = translated_part
            else:
                # Separator was consumed/mangled by the model — fall back to
                # per-cell translation for this row
                logger.debug(
                    "XLSX row separator mismatch: expected %d, got %d; per-cell fallback",
                    len(cells_in_unit), len(parts),
                )
                for cell, orig in cells_in_unit:
                    try:
                        r = translator.translate_text(orig, target_lang)
                        cell.value = r if r is not None else orig
                    except Exception:
                        logger.exception("Per-cell fallback failed")
                        cell.value = orig
                        errors += 1

    # --- write results back (rich-text cells) ---
    # Group translated segments by cell so each cell is written exactly once.
    # cell_rich_updates: id(cell) -> (cell, original_CellRichText, {seg_idx: translated_text})
    cell_rich_updates: dict[int, tuple[Any, CellRichText, dict[int, str]]] = {}
    for (cell, rt, seg_idx, _), tr_text in zip(rich_segs, rich_translated):
        cid = id(cell)
        if cid not in cell_rich_updates:
            cell_rich_updates[cid] = (cell, rt, {})
        if tr_text is not None:
            cell_rich_updates[cid][2][seg_idx] = tr_text

    for cid, (cell, rt, updates) in cell_rich_updates.items():
        if not updates:
            # Every segment translation failed for this cell
            errors += 1
            continue
        new_segs: list[TextBlock | str] = []
        for i, seg in enumerate(rt):
            if i in updates:
                if isinstance(seg, TextBlock):
                    new_segs.append(TextBlock(seg.font, updates[i]))
                else:
                    new_segs.append(updates[i])
            else:
                new_segs.append(seg)
        cell.value = CellRichText(*new_segs)

    # Check for cancellation before saving
    if cancel_event is not None and cancel_event.is_set():
        raise CancelledError("Translation cancelled before saving XLSX")

    wb.save(output_path)
    if errors:
        raise RuntimeError(f"Translation completed with {errors} failed cells")
